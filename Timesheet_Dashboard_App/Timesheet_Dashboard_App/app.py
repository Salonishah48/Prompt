"""
Infinity Globus - Timesheet Efficiency Dashboard Generator
==========================================================
Place your Zoho People timesheet export (.xlsx) in this folder.
The file must contain two sheets:
  1. "Raw Data"
  2. "Expected hours per employee"

Run this script (or double-click run_dashboard.bat) to generate:
  - output/Infinity_Globus_Report.xlsx   (processed Excel with formulas)
  - output/Infinity_Globus_Dashboard.html (interactive 4-tab dashboard)
"""

import os, sys, glob, json
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
TEMPLATE_FILE = os.path.join(SCRIPT_DIR, "dashboard_template.html")

LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
BOLD_FONT = Font(bold=True, name='Arial')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# T1 <-> RM mapping
T1_RM_MAP = {
    "Devendra Singh Nandan Singh Rana": "1275 - Devendra Singh - Rana",
    "Vimal Amrutlal Vadhiya": "1279 - Vimal - Vadhiya",
    "Shah Vivek Rameshbhai": "1001 - Vivek - Shah",
    "Kapil Kanaiyalal Popat": "1249 - Kapil - Popat",
    "Jainam Hiteshkumar Kamdar": "1025 - Jainam - Kamdar",
    "Snehal Vinodkumar Sharma": "1283 - Snehal - Sharma",
    "Prajapati Piyush Dilipbhai": "1039 - Piyush - Prajapati",
    "Vora Parshwa Prakashkumar": "1158 - Parshwa - Vora",
    "Viralkumar Jayantibhai Patel": "1343 - Viralkumar - Patel",
    "Soni Nevil Pramesh": "1020 - Nevil - Soni",
    "Ritesh Jitendrabhai Sanghvi": "224 - Ritesh - Sanghvi",
}

DEPT_COLORS = {
    "Administration": "#64748b", "Business Development": "#0ea5e9",
    "Finance": "#f59e0b", "HR": "#ec4899", "IT": "#06b6d4",
    "Management": "#8b5cf6", "Marketing": "#f97316",
    "UK Accounts and Tax": "#0d9488", "US Accounts": "#1d4ed8",
    "US Audit and Assurance": "#059669", "US Tax": "#7c3aed",
}
DEPT_SHORT = {
    "Administration": "Admin", "Business Development": "BD",
    "UK Accounts and Tax": "UK Accts & Tax", "US Audit and Assurance": "US Audit",
}
CORE_DEPTS = ["US Accounts", "US Audit and Assurance", "US Tax", "UK Accounts and Tax"]


def find_excel_file():
    """Find the first .xlsx/.xlsm file in the script directory."""
    patterns = ["*.xlsx", "*.xlsm"]
    for pat in patterns:
        files = glob.glob(os.path.join(SCRIPT_DIR, pat))
        for f in files:
            fname = os.path.basename(f).lower()
            if "infinity_globus_report" not in fname and "~$" not in fname:
                return f
    return None


def format_header_row(ws, max_col):
    """Apply light green, bold, center, filter, freeze to row 1."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = LIGHT_GREEN
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{ws.max_row}"
    ws.freeze_panes = 'A2'


def process_raw_data(ws):
    """Apply all Raw Data transformations (Steps 2a-2d)."""
    max_row = ws.max_row

    # Step 2a: Move col L values to col M
    for r in range(2, max_row + 1):
        cell_l = ws.cell(row=r, column=12)
        if cell_l.value is not None and cell_l.value != '' and cell_l.value != 0:
            ws.cell(row=r, column=13).value = cell_l.value
            cell_l.value = None

    # Step 2b: Fill empty Billable (M) & Non-Billable (N) with 0
    for r in range(2, max_row + 1):
        for col in [13, 14]:
            cell = ws.cell(row=r, column=col)
            if cell.value is None or cell.value == '':
                cell.value = 0

    # Add new columns P=Leave Hours, Q=Non billable other than leave
    ws.cell(row=1, column=16).value = 'Leave Hours'
    ws.cell(row=1, column=17).value = 'Non billable other than leave'

    # Steps 2c & 2d: Reclassify Timeoffs and Leave/Holiday rows
    for r in range(2, max_row + 1):
        job_name = str(ws.cell(row=r, column=9).value or '').strip()
        work_item = str(ws.cell(row=r, column=10).value or '').strip()
        is_leave = job_name == 'Timeoffs' or work_item in ('Leave', 'Holiday')

        if is_leave:
            total_val = ws.cell(row=r, column=15).value
            ws.cell(row=r, column=16).value = total_val if total_val else 0
            ws.cell(row=r, column=13).value = 0  # Clear Billable
            ws.cell(row=r, column=14).value = 0  # Clear Non-Billable
            ws.cell(row=r, column=17).value = 0
        else:
            ws.cell(row=r, column=16).value = 0
            nb_val = ws.cell(row=r, column=14).value
            ws.cell(row=r, column=17).value = nb_val if nb_val else 0

    # Format row 1
    format_header_row(ws, 17)

    # Column widths
    widths = {1:12, 2:22, 3:26, 4:32, 5:13, 6:32, 7:36, 8:28, 9:38,
              10:28, 11:42, 12:12, 13:11, 14:14, 15:12, 16:13, 17:28}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def build_summary(wb):
    """Create the Summary report sheet with formulas."""
    ws_raw = wb['Raw Data']
    max_row = ws_raw.max_row

    # Gather unique employees
    emp_data = {}
    for r in range(2, max_row + 1):
        eid = ws_raw.cell(row=r, column=5).value
        if eid is None:
            continue
        if eid not in emp_data:
            emp_data[eid] = {
                'dept': ws_raw.cell(row=r, column=2).value or '',
                't1': ws_raw.cell(row=r, column=3).value or '',
                'name': ws_raw.cell(row=r, column=6).value or '',
                'rm': ws_raw.cell(row=r, column=4).value or '',
            }

    ws_sum = wb.create_sheet('Summary report')
    headers = [
        'Department Name', 'T1', 'Employee ID', 'Employee Name',
        'Assigned/Unassigned', 'Billable Hours', 'Non-Billable Hours',
        'Leave Hours', 'Total Hours', 'Expected Hours', 'Effective Hours',
        'Billable Efficiency %', 'Non-Billable Efficiency %',
        'Total Efficiency %', 'Missing Hours'
    ]
    for c, h in enumerate(headers, 1):
        ws_sum.cell(row=1, column=c).value = h

    sorted_emps = sorted(emp_data.items(), key=lambda x: (x[1]['dept'], x[1]['t1'], x[1]['name']))

    row = 2
    for eid, data in sorted_emps:
        ws_sum.cell(row=row, column=1).value = data['dept']
        ws_sum.cell(row=row, column=2).value = data['t1']
        ws_sum.cell(row=row, column=3).value = eid
        ws_sum.cell(row=row, column=4).value = data['name']
        ws_sum.cell(row=row, column=5).value = f"=IFERROR(VLOOKUP(C{row},'Expected hours per employee'!B:F,5,FALSE),\"\")"
        ws_sum.cell(row=row, column=6).value = f"=SUMIFS('Raw Data'!M:M,'Raw Data'!E:E,C{row})"
        ws_sum.cell(row=row, column=7).value = f"=SUMIFS('Raw Data'!N:N,'Raw Data'!E:E,C{row})"
        ws_sum.cell(row=row, column=8).value = f"=SUMIFS('Raw Data'!P:P,'Raw Data'!E:E,C{row})"
        ws_sum.cell(row=row, column=9).value = f"=F{row}+G{row}+H{row}"
        ws_sum.cell(row=row, column=10).value = f"=IFERROR(VLOOKUP(C{row},'Expected hours per employee'!B:E,4,FALSE),0)"
        ws_sum.cell(row=row, column=11).value = f"=J{row}-H{row}"
        ws_sum.cell(row=row, column=12).value = f"=IFERROR(F{row}/K{row},0)"
        ws_sum.cell(row=row, column=12).number_format = '0.00%'
        ws_sum.cell(row=row, column=13).value = f"=IFERROR(G{row}/K{row},0)"
        ws_sum.cell(row=row, column=13).number_format = '0.00%'
        ws_sum.cell(row=row, column=14).value = f"=L{row}+M{row}"
        ws_sum.cell(row=row, column=14).number_format = '0.00%'
        ws_sum.cell(row=row, column=15).value = f"=J{row}-I{row}"
        ws_sum.cell(row=row, column=15).number_format = '0.00'

        for c in [6, 7, 8, 9, 10, 11]:
            ws_sum.cell(row=row, column=c).number_format = '0.00'

        row += 1

    format_header_row(ws_sum, 15)
    widths = {1:22, 2:26, 3:13, 4:32, 5:22, 6:14, 7:16, 8:13,
              9:13, 10:14, 11:14, 12:18, 13:20, 14:16, 15:14}
    for c, w in widths.items():
        ws_sum.column_dimensions[get_column_letter(c)].width = w

    return ws_sum


def extract_dashboard_data(original_excel_path):
    """Read ORIGINAL Excel and extract S (summary) and M (raw) arrays for dashboard.
    Applies all transformation logic in Python (same as openpyxl does for the Excel)."""
    df_raw = pd.read_excel(original_excel_path, 'Raw Data')
    df_exp = pd.read_excel(original_excel_path, 'Expected hours per employee')

    # Build expected hours lookup
    exp_map = {}
    for _, r in df_exp.iterrows():
        eid = r['Employee ID']
        if pd.notna(eid):
            exp_map[eid] = {
                'hours': float(r['Expected Hours']) if pd.notna(r['Expected Hours']) else 0,
                'assigned': str(r['FTE assigned/unassigned']) if pd.notna(r['FTE assigned/unassigned']) else ''
            }

    # Aggregate employee-level data from raw
    emp_agg = {}
    M = []

    for _, r in df_raw.iterrows():
        eid = r['Employee ID']
        if pd.isna(eid):
            continue

        if eid not in emp_agg:
            emp_agg[eid] = {
                'dept': str(r['Department Name']) if pd.notna(r['Department Name']) else '',
                't1': str(r['T1']) if pd.notna(r['T1']) else '',
                'name': str(r['Employee Name']) if pd.notna(r['Employee Name']) else '',
                'rm': str(r['Reporting Manager']) if pd.notna(r['Reporting Manager']) else '',
                'bill': 0, 'nb': 0, 'lv': 0
            }

        # Get raw values
        job_name = str(r.get('Job Name', '')).strip()
        work_item = str(r.get('Work Item', '')).strip()
        total_hours = float(r['Total Hours']) if pd.notna(r.get('Total Hours')) else 0
        is_leave = (job_name == 'Timeoffs') or (work_item in ('Leave', 'Holiday'))

        # Step 2a: Col L ("-No Value-") -> move to Billable
        no_val = r.get('-No Value-', None)
        orig_bill = float(r['Billable']) if pd.notna(r.get('Billable')) else 0
        if pd.notna(no_val) and no_val != '' and no_val != 0:
            orig_bill = float(no_val)

        orig_nb = float(r['Non-Billable']) if pd.notna(r.get('Non-Billable')) else 0

        # Apply leave classification
        if is_leave:
            bill = 0
            nb = 0
            lv = total_hours
            nbotl = 0
            emp_agg[eid]['lv'] += total_hours
        else:
            bill = orig_bill
            nb = orig_nb
            lv = 0
            nbotl = nb
            emp_agg[eid]['bill'] += bill
            emp_agg[eid]['nb'] += nb

        # Date formatting
        date_val = r['Date of Date']
        if pd.notna(date_val):
            if isinstance(date_val, pd.Timestamp):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)
        else:
            date_str = ''

        # Escape any problematic characters in strings for JSON
        def safe_str(v):
            if pd.isna(v) or v is None:
                return ''
            s = str(v)
            return s

        M.append([
            date_str,
            safe_str(r.get('Department Name')),
            safe_str(r.get('T1')),
            safe_str(r.get('Reporting Manager')),
            int(eid) if pd.notna(eid) else 0,
            safe_str(r.get('Employee Name')),
            safe_str(r.get('Client Name')),
            safe_str(r.get('Project Name')),
            job_name, work_item,
            safe_str(r.get('Description')),
            round(bill, 2), round(nb, 2), round(total_hours, 2),
            round(lv, 2), round(nbotl, 2)
        ])

    # Build S array (summary)
    S = []
    for eid, e in emp_agg.items():
        exp = exp_map.get(eid, {'hours': 0, 'assigned': ''})
        tot = e['bill'] + e['nb'] + e['lv']
        eff = exp['hours'] - e['lv']
        be = (e['bill'] / eff * 100) if eff > 0 else 0
        nbe = (e['nb'] / eff * 100) if eff > 0 else 0
        te = be + nbe
        miss = exp['hours'] - tot
        S.append([
            e['dept'], e['t1'], int(eid) if pd.notna(eid) else 0, e['name'], e['rm'],
            round(e['bill'], 2), round(e['nb'], 2), round(e['lv'], 2),
            round(tot, 2), round(exp['hours'], 2), round(eff, 2),
            round(be, 2), round(nbe, 2), round(te, 2), round(miss, 2),
            exp['assigned']
        ])

    dates = [r[0] for r in M if r[0]]
    min_date = min(dates) if dates else ''
    max_date = max(dates) if dates else ''

    return S, M, min_date, max_date


def generate_dashboard(S, M, min_date, max_date, output_path):
    """Generate interactive HTML dashboard from template."""
    template_path = TEMPLATE_FILE
    if not os.path.exists(template_path):
        print(f"  ERROR: dashboard_template.html not found at {template_path}")
        return False

    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()

    html = template.replace('__S_DATA__', json.dumps(S))
    html = html.replace('__M_DATA__', json.dumps(M))
    html = html.replace('__DEPT_COLORS__', json.dumps(DEPT_COLORS))
    html = html.replace('__DEPT_SHORT__', json.dumps(DEPT_SHORT))
    html = html.replace('__CORE_DEPTS__', json.dumps(CORE_DEPTS))
    html = html.replace('__T1_RM_MAP__', json.dumps(T1_RM_MAP))

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    return True


def main():
    print()
    print("=" * 56)
    print("  INFINITY GLOBUS - Timesheet Efficiency Dashboard")
    print("=" * 56)
    print()

    # Step 1: Find Excel file
    xlsx_path = find_excel_file()
    if not xlsx_path:
        print("  ERROR: No .xlsx file found in this folder!")
        print("  Place your Zoho People timesheet export here and try again.")
        print()
        return

    print(f"  Found: {os.path.basename(xlsx_path)}")
    print()

    # Step 2: Process the workbook
    print("  [1/4] Loading workbook...")
    import shutil
    work_path = os.path.join(OUTPUT_DIR, "Infinity_Globus_Report.xlsx")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    shutil.copy(xlsx_path, work_path)

    wb = openpyxl.load_workbook(work_path)
    sheets = wb.sheetnames
    if 'Raw Data' not in sheets:
        print("  ERROR: Sheet 'Raw Data' not found!")
        return
    if 'Expected hours per employee' not in sheets:
        print("  ERROR: Sheet 'Expected hours per employee' not found!")
        return

    # Step 3: Process Raw Data
    print("  [2/4] Processing Raw Data...")
    print("         - Moving Col L values to Billable")
    print("         - Filling empty Billable/Non-Billable with 0")
    print("         - Classifying Timeoffs & Leave/Holiday as Leave Hours")
    process_raw_data(wb['Raw Data'])

    # Format Expected hours sheet
    ws_exp = wb['Expected hours per employee']
    format_header_row(ws_exp, ws_exp.max_column)

    # Step 4: Build Summary Report
    print("  [3/4] Building Summary Report...")
    print("         - Department-wise, T1-wise, Employee-wise aggregation")
    print("         - VLOOKUP formulas for Expected Hours & Assigned/Unassigned")
    print("         - Efficiency calculations (Billable %, Non-Billable %, Total %)")
    print("         - Missing Hours = Expected - Total")

    # Remove existing Summary report if present
    if 'Summary report' in wb.sheetnames:
        del wb['Summary report']
    build_summary(wb)
    wb.save(work_path)

    emp_count = wb['Summary report'].max_row - 1
    print(f"         - {emp_count} employees processed")

    # Step 5: Generate Dashboard
    print("  [4/4] Generating Interactive Dashboard...")
    S, M, min_date, max_date = extract_dashboard_data(xlsx_path)
    dashboard_path = os.path.join(OUTPUT_DIR, "Infinity_Globus_Dashboard.html")
    if generate_dashboard(S, M, min_date, max_date, dashboard_path):
        print(f"         - Date range: {min_date} to {max_date}")
        print(f"         - {len(S)} employees, {len(M)} raw entries")
    else:
        print("         - Dashboard generation failed!")
        return

    # Done
    print()
    print("  " + "=" * 52)
    print("  OUTPUT FILES (in 'output' folder):")
    print(f"    1. Infinity_Globus_Report.xlsx")
    print(f"    2. Infinity_Globus_Dashboard.html")
    print("  " + "=" * 52)
    print()

    # Auto-open dashboard
    print("  Opening dashboard in your browser...")
    import webbrowser
    webbrowser.open(f'file:///{os.path.abspath(dashboard_path)}')
    print()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n  FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        print()
    input("  Press Enter to exit...")
