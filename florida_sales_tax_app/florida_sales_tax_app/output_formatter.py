"""
Output formatter — produces:
  1. A PDF summary of the filled DR-15 return (using reportlab).
  2. An Excel workbook with the same data plus the underlying transactions.

Focused on Line A (regular sales/services) — the only active line in this
configuration.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import List

from tax_engine import DR15Return


# ------------------------------------------------------------------ Excel
def render_excel(
    ret: DR15Return,
    business: dict,
    transactions: List[dict],
    out_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    # ---------- Sheet 1: DR-15 Return ----------
    ws = wb.active
    ws.title = "DR-15 Return"
    info = business["business_info"]

    meta = [
        ("FLORIDA DR-15 SALES AND USE TAX RETURN", ""),
        ("Business Name", info["business_name"]),
        ("Certificate Number", info["certificate_number"]),
        ("FEIN", info["fein"]),
        ("Address", f"{info['physical_address']}, {info['city']}, {info['state']} {info['zip']}"),
        ("County", info["county"]),
        ("Reporting Period", ret.reporting_period),
        ("Surtax Rate", f"{float(ret.surtax_rate)*100:.2f}%"),
        ("", ""),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if i == 1:
            c = ws.cell(row=i, column=1)
            c.font = Font(bold=True, size=14, color="FFFFFF")
            c.fill = header_fill
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        else:
            ws.cell(row=i, column=1).font = bold

    start = len(meta) + 1
    hdr = ["", "Col 1 Gross Sales", "Col 2 Exempt", "Col 3 Taxable", "Col 4 Tax Due"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    r = start + 1
    lines = [
        ("A. Sales/Services/Electricity", ret.line_a),
        ("B. Taxable Purchases (Use Tax)", ret.line_b),
        ("C. Commercial Rentals (repealed 10/1/25)", ret.line_c),
        ("D. Transient Rentals", ret.line_d),
        ("E. Food & Beverage Vending", ret.line_e),
    ]
    for label, line in lines:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=float(line.gross_sales))
        ws.cell(row=r, column=3, value=float(line.exempt_sales))
        ws.cell(row=r, column=4, value=float(line.taxable_amount))
        ws.cell(row=r, column=5, value=float(line.tax_due))
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c >= 2:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = right
        r += 1

    r += 1
    totals = [
        (5, "Total Amount of Tax Due", ret.line_5_total_tax_due, True),
        (6, "Less Lawful Deductions", ret.line_6_lawful_deductions, False),
        (7, "Net Tax Due", ret.line_7_net_tax_due, True),
        (8, "Less Est Tax Paid / DOR Credit Memo", ret.line_8_est_tax_paid_credits, False),
        (9, "Plus Est Tax Due Current Month", ret.line_9_est_tax_due_current, False),
        (10, "Amount Due", ret.line_10_amount_due, True),
        (11, "Less Collection Allowance", ret.line_11_collection_allowance, False),
        (12, "Plus Penalty", ret.line_12_penalty, False),
        (13, "Plus Interest", ret.line_13_interest, False),
        (14, "AMOUNT DUE WITH RETURN", ret.line_14_amount_due_with_return, True),
    ]
    for num, label, val, highlight in totals:
        ws.cell(row=r, column=1, value=f"{num}. {label}").font = bold
        cell = ws.cell(row=r, column=5, value=float(val))
        cell.number_format = '"$"#,##0.00'
        cell.alignment = right
        cell.border = border
        if highlight:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = total_fill
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DISCRETIONARY SALES SURTAX").font = bold
    r += 1
    for num, label, val in [
        ("15(a)", "Exempt Amount of Items Over $5,000", ret.line_15a_exempt_over_5000),
        ("15(b)", "Other Taxable Amounts NOT Subject to Surtax", ret.line_15b_other_not_subject_to_surtax),
        ("15(c)", "Amounts Subject to Surtax at Different Rate", ret.line_15c_different_surtax_rate_amount),
        ("15(d)", "Total Discretionary Sales Surtax Due", ret.line_15d_total_surtax_due),
    ]:
        ws.cell(row=r, column=1, value=f"{num}. {label}")
        cell = ws.cell(row=r, column=5, value=float(val))
        cell.number_format = '"$"#,##0.00'
        cell.alignment = right
        r += 1

    if ret.warnings:
        r += 2
        ws.cell(row=r, column=1, value="NOTES / WARNINGS").font = Font(bold=True, color="C00000")
        r += 1
        for w in ret.warnings:
            ws.cell(row=r, column=1, value=f"- {w}")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1

    # Reviewer reference
    r += 2
    ref_cell = ws.cell(row=r, column=1, value="FOR REVIEWERS / AUDITORS")
    ref_cell.font = Font(bold=True, color="1F4E78")
    r += 1
    for text in [
        "• Every Line A total above is derived from the 'Line-by-Line Working' "
        "sheet — see the TOTALS row at the bottom of that sheet.",
        "• Each transaction's tax computation is shown individually with rates, "
        "surtax base (showing when the $5k cap was applied), and the per-row "
        "gap between calculated and Shopify-collected tax.",
        "• The 'Methodology' sheet documents every formula and decision rule "
        "used to produce these numbers.",
    ]:
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30
        r += 1

    for i, w in enumerate([50, 18, 18, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Sheet 2: Transactions (audit-grade line-by-line working) ----------
    ws2 = wb.create_sheet("Line-by-Line Working")
    has_shopify = any("order_number" in t for t in transactions)

    # Top banner explaining the sheet
    banner = ws2.cell(row=1, column=1,
                      value="LINE-BY-LINE TAX CALCULATION — AUDIT TRAIL")
    banner.font = Font(bold=True, size=13, color="FFFFFF")
    banner.fill = header_fill
    banner.alignment = Alignment(horizontal="center")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)

    note = ws2.cell(row=2, column=1,
        value=("Each row = one Shopify line item (after merging Shopify's "
               "State+County jurisdiction splits). Columns show both what "
               "Shopify charged and what Florida law requires, with the gap "
               "for each transaction."))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    note.font = Font(italic=True, color="666666")
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=17)
    ws2.row_dimensions[2].height = 30

    if has_shopify:
        headers2 = [
            "#", "Date", "Order #", "Channel",
            "Col 1 Gross", "Col 3 Taxable", "Exempt (Col1-Col3)",
            "Dest City", "County",
            "State Rate", "State Tax (calc)",
            "Surtax Rate", "Surtax Base", "Surtax (calc)", "$5k Cap Applied?",
            "Total Tax (calc)", "Shopify Tax", "Gap (calc - Shopify)"
        ]
    else:
        headers2 = ["#", "Date", "Amount", "Exempt", "Delivery County",
                    "Single TPP Item", "Source File"]

    header_row_num = 4
    for c, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=header_row_num, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    ws2.row_dimensions[header_row_num].height = 30

    # Data rows
    data_start = header_row_num + 1
    for i, t in enumerate(transactions, start=1):
        r_num = data_start + i - 1
        if has_shopify:
            gross_v = t.get("gross_sales", 0)
            taxable_v = t.get("taxable_amount", 0)
            exempt_portion = gross_v - taxable_v
            ws2.cell(row=r_num, column=1, value=i)
            ws2.cell(row=r_num, column=2, value=t.get("date", ""))
            ws2.cell(row=r_num, column=3, value=t.get("order_number", ""))
            ws2.cell(row=r_num, column=4, value=t.get("channel", ""))
            ws2.cell(row=r_num, column=5, value=gross_v).number_format = '"$"#,##0.00'
            ws2.cell(row=r_num, column=6, value=taxable_v).number_format = '"$"#,##0.00'
            ws2.cell(row=r_num, column=7, value=exempt_portion).number_format = '"$"#,##0.00'
            ws2.cell(row=r_num, column=8, value=t.get("destination_city", ""))
            ws2.cell(row=r_num, column=9,
                     value=(t.get("delivery_county") or "").replace("_", " "))
            # Calculated values (only present if Florida sale and computed)
            if "_calc_state_rate" in t:
                ws2.cell(row=r_num, column=10,
                         value=t["_calc_state_rate"]).number_format = '0.00%'
                ws2.cell(row=r_num, column=11,
                         value=t["_calc_state_tax"]).number_format = '"$"#,##0.00'
                ws2.cell(row=r_num, column=12,
                         value=t["_calc_surtax_rate"]).number_format = '0.00%'
                ws2.cell(row=r_num, column=13,
                         value=t["_calc_surtax_base"]).number_format = '"$"#,##0.00'
                ws2.cell(row=r_num, column=14,
                         value=t["_calc_surtax"]).number_format = '"$"#,##0.00'
                ws2.cell(row=r_num, column=15,
                         value="Yes" if t.get("_calc_cap_applied") else "No")
                ws2.cell(row=r_num, column=16,
                         value=t["_calc_total_tax"]).number_format = '"$"#,##0.00'
                gap_v = t.get("_calc_gap", 0)
                gap_cell = ws2.cell(row=r_num, column=18, value=gap_v)
                gap_cell.number_format = '"$"#,##0.00'
                # Red if gap > $0.05
                if abs(gap_v) > 0.05:
                    gap_cell.fill = PatternFill("solid", fgColor="FFD6D6")
            else:
                ws2.cell(row=r_num, column=10, value="(skipped — not FL)")
            ws2.cell(row=r_num, column=17,
                     value=t.get("shopify_tax_amount", 0)).number_format = '"$"#,##0.00'
        else:
            ws2.cell(row=r_num, column=1, value=i)
            ws2.cell(row=r_num, column=2, value=t.get("date", ""))
            amt_cell = ws2.cell(row=r_num, column=3, value=t.get("amount", 0))
            amt_cell.number_format = '"$"#,##0.00'
            ws2.cell(row=r_num, column=4, value="Yes" if t.get("exempt") else "No")
            ws2.cell(row=r_num, column=5,
                     value=t.get("delivery_county") or business["business_info"]["county"])
            ws2.cell(row=r_num, column=6,
                     value="Yes" if t.get("is_single_tpp_item", True) else "No")
            ws2.cell(row=r_num, column=7, value=t.get("_source_file", ""))

    # Totals row at the bottom
    if has_shopify and transactions:
        totals_row = data_start + len(transactions) + 1
        ws2.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
        ws2.merge_cells(start_row=totals_row, start_column=1,
                        end_row=totals_row, end_column=4)
        for col_idx, field in [(5, "gross_sales"), (6, "taxable_amount")]:
            cell = ws2.cell(row=totals_row, column=col_idx,
                            value=f"=SUM({get_column_letter(col_idx)}{data_start}:"
                                  f"{get_column_letter(col_idx)}{totals_row-2})")
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
        # Col 7 exempt = col5 - col6
        cell = ws2.cell(row=totals_row, column=7,
                        value=f"=E{totals_row}-F{totals_row}")
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        # Sum calculated tax, shopify tax, gap
        for col_idx in [11, 14, 16, 17, 18]:
            cell = ws2.cell(row=totals_row, column=col_idx,
                            value=f"=SUM({get_column_letter(col_idx)}{data_start}:"
                                  f"{get_column_letter(col_idx)}{totals_row-2})")
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

    if has_shopify:
        widths2 = [5, 12, 10, 14, 12, 12, 14, 14, 14, 10, 12, 10, 12, 12, 12, 14, 12, 14]
    else:
        widths2 = [5, 12, 14, 10, 18, 16, 28]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = ws2.cell(row=data_start, column=1)

    # ---------- Sheet 3: Methodology (how we computed everything) ----------
    ws3 = wb.create_sheet("Methodology")

    wrap_left = Alignment(wrap_text=True, vertical="top")
    h1_style = Font(bold=True, size=14, color="1F4E78")
    h2_style = Font(bold=True, size=11, color="1F4E78")
    mono = Font(name="Consolas", size=10)

    methodology = [
        ("h1", "How This DR-15 Was Calculated"),
        ("p",  "This workbook is generated from a Shopify 'Sales by tax' export. "
               "The logic below documents every step so any reviewer, bookkeeper, "
               "or auditor can trace each number back to its source."),
        ("", ""),

        ("h2", "1. Loading and de-duplicating Shopify rows"),
        ("p",  "Shopify exports ONE ROW PER TAX JURISDICTION. For a Florida sale "
               "in a surtax county, each line item produces 2 rows: one 'State' "
               "row (6% tax) and one 'County' row (county surtax). The app merges "
               "these pairs back into ONE transaction per line item so gross "
               "sales are not double-counted."),
        ("p",  "Grouping key: (order_number, line_item_id, date, taxable_amount, "
               "gross_sales). Using taxable + gross as part of the key correctly "
               "splits cases where one line_item_id has multiple distinct amounts "
               "(e.g. the same SKU sold at different prices)."),
        ("p",  f"In this run: {len(transactions)} unique line items were produced "
               f"after de-duplication."),
        ("", ""),

        ("h2", "2. Filtering out non-real rows"),
        ("p",  "Two kinds of rows are skipped automatically:"),
        ("p",  "  • Shopify 'ghost' / subtotal rows — rows with no order number "
               "AND no destination city (Shopify sometimes includes aggregates "
               "at the bottom of the export)."),
        ("p",  "  • Non-Florida sales — rows where destination state is not "
               "Florida. Florida has no jurisdiction over these unless you have "
               "economic nexus in the destination state."),
        ("", ""),

        ("h2", "3. DR-15 Column 1 — Gross Sales"),
        ("p",  "Formula: sum of Shopify's 'Gross sales on line items' column "
               "across all included line items."),
        ("p",  f"This run: ${float(ret.line_a.gross_sales):,.2f}"),
        ("", ""),

        ("h2", "4. DR-15 Column 3 — Taxable Amount"),
        ("p",  "Formula: sum of Shopify's 'Taxable amount' column across all "
               "included line items."),
        ("p",  "We use Shopify's column directly rather than computing it "
               "ourselves because Shopify is the authoritative record of what "
               "was actually deemed taxable at checkout. FDOR will reconcile "
               "against this number on audit."),
        ("p",  f"This run: ${float(ret.line_a.taxable_amount):,.2f}"),
        ("", ""),

        ("h2", "5. DR-15 Column 2 — Exempt Sales"),
        ("p",  "Formula: Col 1 − Col 3 (Gross − Taxable)."),
        ("p",  "This single number correctly absorbs:"),
        ("p",  "  • Discounts (reduce taxable but keep gross)"),
        ("p",  "  • Resale exemptions (customer had a DR-13)"),
        ("p",  "  • Non-taxable categories"),
        ("p",  "  • Return/refund credits (negative taxable, positive gross)"),
        ("p",  "  • Out-of-state portions not subject to Florida tax"),
        ("p",  "We chose this formula over summing Shopify's 'Exempt amount' "
               "column because that column alone doesn't capture discounts, "
               "shipping, or return offsets."),
        ("p",  f"This run: ${float(ret.line_a.gross_sales):,.2f} − "
               f"${float(ret.line_a.taxable_amount):,.2f} = "
               f"${float(ret.line_a.exempt_sales):,.2f}"),
        ("", ""),

        ("h2", "6. DR-15 Column 4 — Tax Due (per line item)"),
        ("p",  "For EACH line item, the app independently computes:"),
        ("p",  "  state_tax = taxable × 6%"),
        ("p",  "  delivery_surtax_rate = rate for county where item was delivered"),
        ("p",  "  if (single TPP item AND taxable > $5,000):"),
        ("p",  "      surtax_base = $5,000   ← Florida's single-item cap"),
        ("p",  "  else:"),
        ("p",  "      surtax_base = taxable"),
        ("p",  "  surtax = surtax_base × delivery_surtax_rate"),
        ("p",  "  total_tax = state_tax + surtax"),
        ("p",  "Rounding follows Florida's DR-15N rule: carry to 3 decimals, "
               "round UP to next cent when third decimal > 4."),
        ("", ""),

        ("h2", "7. Line 15(a) — Amount of Items Over $5,000"),
        ("p",  "Tracks the portion of each single-TPP-item sale that exceeded "
               "$5,000 and therefore was excluded from surtax. "
               "For example, an $8,500 single item adds $3,500 to Line 15(a) "
               "(the portion above the $5k cap)."),
        ("p",  f"This run: ${float(ret.line_15a_exempt_over_5000):,.2f}"),
        ("", ""),

        ("h2", "8. Line 15(c) — Amounts at a Different Surtax Rate"),
        ("p",  f"Any line item whose delivery county has a different surtax rate "
               f"than your business county ({business['business_info']['county']}, "
               f"{float(ret.surtax_rate)*100:.2f}%) is reported here. This is "
               "required because out-of-county deliveries use destination-based "
               "surtax per FL rule 12A-15.008."),
        ("p",  f"This run: ${float(ret.line_15c_different_surtax_rate_amount):,.2f}"),
        ("", ""),

        ("h2", "9. Line 15(d) — Total Discretionary Sales Surtax Due"),
        ("p",  "Sum of per-transaction surtax amounts computed in step 6. "
               "This is the county portion only (state tax reported separately)."),
        ("p",  f"This run: ${float(ret.line_15d_total_surtax_due):,.2f}"),
        ("", ""),

        ("h2", "10. Line 11 — Collection Allowance"),
        ("p",  "Florida grants a 2.5% collection allowance on the first $1,200 "
               "of tax due, capped at $30. Only available if return is filed "
               "AND paid electronically AND on time."),
        ("p",  f"This run: ${float(ret.line_11_collection_allowance):,.2f}"),
        ("", ""),

        ("h2", "11. Shopify Collection Reconciliation"),
        ("p",  "Two separate numbers are tracked:"),
        ("p",  "  • Tax calculated by app = what Florida law requires"),
        ("p",  "  • Tax collected by Shopify = sum of Shopify's 'Tax amount' column"),
        ("p",  "Any gap means Shopify's configuration has drifted from the "
               "correct rates, or the $5k cap is being applied incorrectly "
               "(Shopify sometimes applies the cap per-order instead of "
               "per-line-item). You must remit the FULL calculated amount "
               "and absorb any shortfall."),
        ("p",  f"This run: calculated ${float(ret.line_a.tax_due):,.2f} vs "
               f"Shopify ${float(ret.shopify_tax_collected):,.2f}, gap "
               f"${float(ret.tax_gap):,.2f}"),
        ("", ""),

        ("h2", "Reference to output file"),
        ("p",  "• Sheet 'DR-15 Return' — the final filled form (matches the PDF)."),
        ("p",  "• Sheet 'Line-by-Line Working' — every transaction with its "
               "own tax computation, per-row gap vs Shopify, and totals row."),
        ("p",  "• Sheet 'Methodology' — this sheet."),
        ("", ""),
        ("p",  "For audit defense: every number on Sheet 1 can be traced to "
               "summed columns on Sheet 2. Every computed value on Sheet 2 can "
               "be re-derived from the formulas documented here."),
    ]

    r = 1
    for kind, text in methodology:
        cell = ws3.cell(row=r, column=1, value=text)
        cell.alignment = wrap_left
        if kind == "h1":
            cell.font = h1_style
            ws3.row_dimensions[r].height = 22
        elif kind == "h2":
            cell.font = h2_style
            ws3.row_dimensions[r].height = 18
        else:
            # Paragraph rows — calculate row height based on text length
            text_len = len(text) if text else 0
            if text_len > 100:
                ws3.row_dimensions[r].height = 30
            elif text_len > 60:
                ws3.row_dimensions[r].height = 22
            if text.startswith("  •") or text.startswith("  "):
                cell.font = mono
        r += 1

    ws3.column_dimensions["A"].width = 110
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    wb.save(out_path)


# ------------------------------------------------------------------ PDF
def render_pdf(
    ret: DR15Return,
    business: dict,
    transactions: List[dict],
    out_path: Path,
) -> None:
    """Render a polished PDF summary mirroring the DR-15 layout."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        KeepTogether,
    )

    info = business["business_info"]
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1F4E78"),
        alignment=1, spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=10, alignment=1, textColor=colors.grey, spaceAfter=14,
    )
    h2 = ParagraphStyle(
        "H2", parent=styles["Heading3"],
        fontSize=11, textColor=colors.HexColor("#1F4E78"), spaceBefore=10, spaceAfter=4,
    )
    note = ParagraphStyle(
        "Note", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#C00000"),
    )

    story = []

    story.append(Paragraph("Florida DR-15 Sales and Use Tax Return", title_style))
    story.append(Paragraph(
        f"Filled summary for reporting period <b>{ret.reporting_period}</b>",
        subtitle_style,
    ))

    # Business info table
    biz_rows = [
        ["Business Name", info["business_name"]],
        ["Certificate Number", info["certificate_number"]],
        ["FEIN", info["fein"]],
        ["Address", f"{info['physical_address']}, {info['city']}, {info['state']} {info['zip']}"],
        ["County", info["county"]],
        ["Surtax Rate", f"{float(ret.surtax_rate)*100:.2f}%"],
        ["Reporting Period", ret.reporting_period],
        ["Filing Mode", "Electronic" if business["filing_preferences"]["files_electronically"] else "Paper"],
        ["Late Filing?", "Yes" if ret.late_filing else "No"],
    ]
    biz_tbl = Table(biz_rows, colWidths=[2.0 * inch, 4.3 * inch])
    biz_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F2F2F2")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(biz_tbl)

    # Front of form
    story.append(Paragraph("Front of Return — Lines A through E", h2))
    money = lambda d: f"${float(d):,.2f}"

    front = [
        ["", "Col 1 Gross Sales", "Col 2 Exempt", "Col 3 Taxable", "Col 4 Tax Due"],
        ["A. Sales / Services / Electricity",
         money(ret.line_a.gross_sales), money(ret.line_a.exempt_sales),
         money(ret.line_a.taxable_amount), money(ret.line_a.tax_due)],
        ["B. Taxable Purchases (Use Tax)",
         money(ret.line_b.gross_sales), money(ret.line_b.exempt_sales),
         money(ret.line_b.taxable_amount), money(ret.line_b.tax_due)],
        ["C. Commercial Rentals (repealed 10/1/25)",
         money(ret.line_c.gross_sales), money(ret.line_c.exempt_sales),
         money(ret.line_c.taxable_amount), money(ret.line_c.tax_due)],
        ["D. Transient Rentals",
         money(ret.line_d.gross_sales), money(ret.line_d.exempt_sales),
         money(ret.line_d.taxable_amount), money(ret.line_d.tax_due)],
        ["E. Food & Beverage Vending",
         money(ret.line_e.gross_sales), money(ret.line_e.exempt_sales),
         money(ret.line_e.taxable_amount), money(ret.line_e.tax_due)],
    ]
    front_tbl = Table(front, colWidths=[2.4 * inch, 1.05 * inch, 0.95 * inch, 1.05 * inch, 0.95 * inch])
    front_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(front_tbl)

    # Totals block (Lines 5-14)
    story.append(Paragraph("Tax Computation — Lines 5 through 14", h2))
    totals_rows = [
        ["5. Total Amount of Tax Due", money(ret.line_5_total_tax_due)],
        ["6. Less Lawful Deductions", money(ret.line_6_lawful_deductions)],
        ["7. Net Tax Due", money(ret.line_7_net_tax_due)],
        ["8. Less Est Tax Paid / DOR Credit Memo", money(ret.line_8_est_tax_paid_credits)],
        ["9. Plus Est Tax Due Current Month", money(ret.line_9_est_tax_due_current)],
        ["10. Amount Due", money(ret.line_10_amount_due)],
        ["11. Less Collection Allowance", money(ret.line_11_collection_allowance)],
        ["12. Plus Penalty", money(ret.line_12_penalty)],
        ["13. Plus Interest", money(ret.line_13_interest)],
        ["14. AMOUNT DUE WITH RETURN", money(ret.line_14_amount_due_with_return)],
    ]
    totals_tbl = Table(totals_rows, colWidths=[4.4 * inch, 1.9 * inch])
    totals_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),   # Line 7
        ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),   # Line 10
        ("FONTNAME", (0, 9), (-1, 9), "Helvetica-Bold"),   # Line 14
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF2CC")),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#FFF2CC")),
        ("BACKGROUND", (0, 5), (-1, 5), colors.HexColor("#FFF2CC")),
        ("BACKGROUND", (0, 9), (-1, 9), colors.HexColor("#FFE699")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(totals_tbl)

    # Surtax detail — keep heading + table together so they don't split
    surtax_heading = Paragraph("Discretionary Sales Surtax — Lines 15(a) through 15(d)", h2)
    surtax_rows = [
        ["15(a). Exempt Amount of Items Over $5,000", money(ret.line_15a_exempt_over_5000)],
        ["15(b). Other Taxable Amounts NOT Subject to Surtax", money(ret.line_15b_other_not_subject_to_surtax)],
        ["15(c). Amounts Subject to Surtax at Different Rate", money(ret.line_15c_different_surtax_rate_amount)],
        ["15(d). Total Discretionary Sales Surtax Due", money(ret.line_15d_total_surtax_due)],
    ]
    s_tbl = Table(surtax_rows, colWidths=[4.4 * inch, 1.9 * inch])
    s_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
        ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#FFE699")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(KeepTogether([surtax_heading, s_tbl]))

    # Shopify reconciliation (only shown if Shopify data was used)
    if float(ret.shopify_tax_collected) > 0:
        recon_heading = Paragraph("Shopify Collection Reconciliation", h2)
        gap = float(ret.tax_gap)
        gap_label = ("Under-collection (must absorb)" if gap > 0
                     else ("Over-collection (must remit)" if gap < 0 else "Match"))
        gap_color = colors.HexColor("#FFCCCC") if gap > 0 else (
            colors.HexColor("#CCE5FF") if gap < 0 else colors.HexColor("#D4EDDA"))
        recon_rows = [
            ["Tax calculated by app (state + county surtax)", money(ret.line_a.tax_due)],
            ["Tax collected by Shopify", money(ret.shopify_tax_collected)],
            [f"Gap — {gap_label}", money(abs(Decimal(str(gap))))],
        ]
        recon_tbl = Table(recon_rows, colWidths=[4.4 * inch, 1.9 * inch])
        recon_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
            ("BACKGROUND", (0, 2), (-1, 2), gap_color),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(KeepTogether([recon_heading, recon_tbl]))

    # Warnings / Notes
    if ret.warnings:
        story.append(Paragraph("Notes / Warnings", h2))
        for w in ret.warnings:
            story.append(Paragraph("• " + w, note))

    # Footer note
    story.append(Spacer(1, 0.2 * inch))
    footer = ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontSize=8, textColor=colors.grey, alignment=1,
    )
    story.append(Paragraph(
        "This PDF is a preparation aid. File the return at "
        "<b>floridarevenue.com/taxes/eServices</b>. "
        "Returns are due on the 1st of the month following the reporting period "
        "and are late after the 20th.",
        footer,
    ))

    # Page 2: Transaction detail
    story.append(PageBreak())
    story.append(Paragraph("Supporting Transactions", title_style))
    story.append(Paragraph(
        f"{len(transactions)} transactions loaded for period {ret.reporting_period}",
        subtitle_style,
    ))

    if transactions:
        # Detect if we have Shopify data
        has_shopify = any("order_number" in t for t in transactions)

        if has_shopify:
            header_row = ["Date", "Order #", "Gross", "Taxable",
                          "Shopify Tax", "Dest City", "County"]
            data = [header_row]
            for t in transactions:
                data.append([
                    t.get("date", ""),
                    t.get("order_number", ""),
                    money(t.get("gross_sales", 0)),
                    money(t.get("taxable_amount", 0)),
                    money(t.get("shopify_tax_amount", 0)),
                    t.get("destination_city", ""),
                    (t.get("delivery_county") or "").replace("_", " "),
                ])
            col_widths = [0.85 * inch, 0.8 * inch, 0.85 * inch,
                          0.85 * inch, 0.95 * inch, 1.2 * inch, 1.1 * inch]
        else:
            header_row = ["Date", "Amount", "Exempt", "Delivery County",
                          "Single TPP?", "Source"]
            data = [header_row]
            for t in transactions:
                data.append([
                    t.get("date", ""),
                    money(t.get("amount", 0)),
                    "Yes" if t.get("exempt") else "No",
                    t.get("delivery_county") or info["county"],
                    "Yes" if t.get("is_single_tpp_item", True) else "No",
                    t.get("_source_file", ""),
                ])
            col_widths = [0.95 * inch, 1.0 * inch, 0.7 * inch,
                          1.3 * inch, 0.9 * inch, 1.55 * inch]

        tbl_style = [
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F7F7F7")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        if has_shopify:
            tbl_style.append(("ALIGN", (2, 1), (4, -1), "RIGHT"))
        else:
            tbl_style.append(("ALIGN", (1, 1), (1, -1), "RIGHT"))

        txn_tbl = Table(data, colWidths=col_widths, repeatRows=1)
        txn_tbl.setStyle(TableStyle(tbl_style))
        story.append(txn_tbl)
    else:
        story.append(Paragraph("No transactions found for this period.", styles["Normal"]))

    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        title=f"DR-15 {ret.reporting_period}",
        author=info["business_name"],
    )
    doc.build(story)
