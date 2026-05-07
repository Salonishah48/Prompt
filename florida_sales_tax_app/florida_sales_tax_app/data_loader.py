"""
Daily data loader — Excel / CSV files in Shopify tax export format.

Handles the 38-column Shopify export:
  Is filed by channel, Channel, Order number, Line item ID, Sale date,
  Gross sales on line items, Discounts, Returns, Net sales on line items,
  Shipping, Exempt amount, Non-taxable amount, Non-taxed amount,
  Taxable amount, Tax rate, Tax amount, Tax jurisdiction type,
  Tax jurisdiction, Tax county, Tax jurisdiction code,
  Destination country/state/city/address/zip, Billing ..., Origin ...,
  Product category code, Tax exemptions, Shopify reference ID.

Each Shopify line-item row is converted to one engine-friendly transaction:
  {
    "date": "YYYY-MM-DD",
    "order_number": "#1038",
    "gross_sales": 400.00,           # Gross sales on line items
    "discounts": 28.00,              # abs value of Discounts column
    "returns": 0.00,
    "net_sales": 372.00,             # Net sales on line items
    "shipping": 0.00,
    "exempt_amount": 0.00,
    "taxable_amount": 372.00,
    "shopify_tax_rate": 0.06,
    "shopify_tax_amount": 22.32,     # what Shopify actually collected
    "delivery_county": "Hillsborough",    # inferred from dest city/zip
    "destination_city": "Plant City",
    "destination_state": "Florida",
    "destination_zip": "33566-4744",
    "product_category": "Home & Garden > ...",
    "is_single_tpp_item": True,       # default; services/rentals would be False
    "_source_file": "july_2025.xlsx",
  }

The tax engine then recomputes expected tax and compares to Shopify's.
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, date
from typing import List, Optional, Tuple


# Canonical Shopify headers, all lowercased for matching
SHOPIFY_HEADERS = [
    "is filed by channel", "channel", "order number", "line item id", "sale date",
    "gross sales on line items", "discounts", "returns", "net sales on line items",
    "shipping", "exempt amount", "non-taxable amount", "non-taxed amount",
    "taxable amount", "tax rate", "tax amount", "tax jurisdiction type",
    "tax jurisdiction", "tax county", "tax jurisdiction code",
    "destination country", "destination state", "destination city",
    "destination address", "destination zip",
    "billing address", "billing city", "billing state", "billing country", "billing zip",
    "origin address", "origin city", "origin state", "origin country", "origin zip",
    "product category code", "tax exemptions", "shopify reference id",
]


# Mapping of common FL city -> county for cases where Shopify leaves the
# county blank. Extend this as needed; the app falls back to the business
# county from the config when neither the export nor this table has it.
FL_CITY_TO_COUNTY = {
    # Hillsborough
    "tampa": "Hillsborough", "plant city": "Hillsborough",
    "brandon": "Hillsborough", "riverview": "Hillsborough",
    "valrico": "Hillsborough", "lutz": "Hillsborough",
    "apollo beach": "Hillsborough", "ruskin": "Hillsborough",
    "sun city center": "Hillsborough", "wimauma": "Hillsborough",
    "gibsonton": "Hillsborough", "seffner": "Hillsborough",
    "dover": "Hillsborough", "thonotosassa": "Hillsborough",
    # Pinellas
    "st petersburg": "Pinellas", "st. petersburg": "Pinellas",
    "saint petersburg": "Pinellas", "clearwater": "Pinellas",
    "largo": "Pinellas", "palm harbor": "Pinellas",
    "dunedin": "Pinellas", "pinellas park": "Pinellas",
    "tarpon springs": "Pinellas", "seminole": "Pinellas",
    # Orange
    "orlando": "Orange", "winter park": "Orange", "ocoee": "Orange",
    "apopka": "Orange", "winter garden": "Orange",
    # Miami-Dade
    "miami": "Miami_Dade", "miami beach": "Miami_Dade",
    "hialeah": "Miami_Dade", "doral": "Miami_Dade", "homestead": "Miami_Dade",
    # Broward
    "fort lauderdale": "Broward", "hollywood": "Broward",
    "pembroke pines": "Broward", "coral springs": "Broward",
    # Duval
    "jacksonville": "Duval", "jacksonville beach": "Duval",
    # Palm Beach
    "west palm beach": "Palm_Beach", "boca raton": "Palm_Beach",
    "boynton beach": "Palm_Beach", "delray beach": "Palm_Beach",
    # Lee
    "fort myers": "Lee", "cape coral": "Lee", "estero": "Lee",
    # Leon
    "tallahassee": "Leon",
    # Sarasota
    "sarasota": "Sarasota", "venice": "Sarasota",
    # Polk
    "lakeland": "Polk", "winter haven": "Polk",
    # Alachua
    "gainesville": "Alachua",
    # Seminole (the county, not the city above)
    "sanford": "Seminole", "altamonte springs": "Seminole",
    # Osceola
    "kissimmee": "Osceola", "st cloud": "Osceola",
    # Volusia
    "daytona beach": "Volusia", "deland": "Volusia",
    # Monroe
    "key west": "Monroe", "marathon": "Monroe",
    # Collier
    "naples": "Collier",
    # Brevard
    "melbourne": "Brevard", "palm bay": "Brevard", "titusville": "Brevard",
    # Charlotte
    "port charlotte": "Charlotte", "punta gorda": "Charlotte",
    # Manatee
    "bradenton": "Manatee", "palmetto": "Manatee", "ellenton": "Manatee",
    # Pasco
    "new port richey": "Pasco", "wesley chapel": "Pasco",
    "land o lakes": "Pasco", "port richey": "Pasco",
    # St Johns
    "st augustine": "St_Johns", "saint augustine": "St_Johns",
    "ponte vedra": "St_Johns", "ponte vedra beach": "St_Johns",
    # Indian River
    "vero beach": "Indian_River",
    # St Lucie
    "port st lucie": "St_Lucie", "port saint lucie": "St_Lucie",
    "fort pierce": "St_Lucie",
    # Martin
    "stuart": "Martin", "jensen beach": "Martin",
    # Marion
    "ocala": "Marion",
    # Lake
    "clermont": "Lake", "leesburg": "Lake", "eustis": "Lake",
    # Santa Rosa
    "milton": "Santa_Rosa", "pace": "Santa_Rosa",
    # Escambia
    "pensacola": "Escambia",
    # Okaloosa
    "destin": "Okaloosa", "fort walton beach": "Okaloosa", "crestview": "Okaloosa",
    # Walton
    "defuniak springs": "Walton", "santa rosa beach": "Walton", "miramar beach": "Walton",
    # Bay
    "panama city": "Bay", "panama city beach": "Bay",
}


def _as_float(val, default=0.0) -> float:
    if val is None or val == "" or val == "-" or (isinstance(val, str) and val.strip() == "-"):
        return default
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def _as_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s == "-" else s


def _as_iso_date(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s[:10]).strftime("%Y-%m-%d")
    except ValueError:
        return s


def _infer_county(row: dict) -> Optional[str]:
    """Figure out the delivery county from the Shopify row."""
    # 1. Direct 'Tax county' column if populated
    county = _as_str(row.get("tax county"))
    if county:
        return county.replace(" ", "_").replace("-", "_")
    # 2. City lookup table
    city = _as_str(row.get("destination city")).lower()
    if city and city in FL_CITY_TO_COUNTY:
        return FL_CITY_TO_COUNTY[city]
    # 3. Billing city fallback
    bill_city = _as_str(row.get("billing city")).lower()
    if bill_city and bill_city in FL_CITY_TO_COUNTY:
        return FL_CITY_TO_COUNTY[bill_city]
    return None


def _looks_like_service(category: str) -> bool:
    """Heuristic: services / rentals don't get the $5,000 surtax cap."""
    c = (category or "").lower()
    service_hints = ["service", "rental", "admission", "subscription"]
    return any(h in c for h in service_hints)


def _normalize_headers(raw_headers) -> List[str]:
    return [str(h).strip().lower() if h is not None else "" for h in raw_headers]


def _row_to_txn(row: dict, source_file: str) -> Optional[dict]:
    """Convert one Shopify row to an intermediate representation.
    Each row represents ONE jurisdiction's tax on a line item — we need to
    aggregate pairs (State + County) later before passing to the engine."""
    gross = _as_float(row.get("gross sales on line items"))
    net = _as_float(row.get("net sales on line items"))
    taxable = _as_float(row.get("taxable amount"))
    date_str = _as_iso_date(row.get("sale date"))
    tax_amount = _as_float(row.get("tax amount"))
    order_number = _as_str(row.get("order number"))

    # Skip truly blank rows (no data at all)
    if gross == 0 and net == 0 and taxable == 0 and tax_amount == 0 and not date_str:
        return None

    # Skip Shopify "subtotal" / "summary" ghost rows — they have no order
    # number AND no destination city (they're aggregates, not real sales).
    dest_city_check = _as_str(row.get("destination city"))
    if not order_number and not dest_city_check:
        return None

    discounts = abs(_as_float(row.get("discounts")))
    returns = abs(_as_float(row.get("returns")))
    exempt = _as_float(row.get("exempt amount"))
    non_taxable = _as_float(row.get("non-taxable amount"))
    non_taxed = _as_float(row.get("non-taxed amount"))

    county = _infer_county(row)
    dest_state = _as_str(row.get("destination state"))
    dest_country = _as_str(row.get("destination country"))

    is_florida_sale = (dest_state.lower() == "florida" or
                       (dest_country.lower() == "united states" and not dest_state))

    category = _as_str(row.get("product category code"))
    juris_type = _as_str(row.get("tax jurisdiction type")).lower()

    return {
        "date": date_str,
        "order_number": order_number,
        "line_item_id": _as_str(row.get("line item id")),
        "channel": _as_str(row.get("channel")),
        # Sales breakdown (same on both State and County rows for a given line item)
        "gross_sales": gross,
        "discounts": discounts,
        "returns": returns,
        "net_sales": net,
        "shipping": _as_float(row.get("shipping")),
        "exempt_amount": exempt + non_taxable + non_taxed,
        "taxable_amount": taxable,
        # This row's jurisdiction-specific tax
        "jurisdiction_type": juris_type,   # 'state' or 'county'
        "shopify_tax_rate": _as_float(row.get("tax rate")),
        "shopify_tax_amount": tax_amount,
        # Engine-facing fields
        "amount": taxable,
        "exempt": False,
        "delivery_county": county,
        "destination_city": _as_str(row.get("destination city")),
        "destination_state": dest_state,
        "destination_zip": _as_str(row.get("destination zip")),
        "is_single_tpp_item": not _looks_like_service(category),
        "product_category": category,
        "is_florida_sale": is_florida_sale,
        "_source_file": source_file,
    }


def _aggregate_shopify_rows(raw_rows: List[dict]) -> List[dict]:
    """
    Shopify exports ONE ROW PER TAX JURISDICTION. For a Florida sale in a
    surtax county, each line item produces:
      - One 'State' row  (6% state tax on gross/taxable)
      - One 'County' row (1.5% surtax on the same gross/taxable)

    This function collapses those pairs back into ONE transaction per
    unique (order, line_item, taxable_amount) so we don't double-count
    gross sales. The Shopify tax collected is the SUM of both rows.

    Key used for grouping: (order_number, line_item_id, date,
    taxable_amount, gross_sales). Using taxable_amount + gross_sales as
    part of the key correctly splits cases where one line_item_id has
    multiple distinct amounts (as in order #1039 with 3 different
    quantities).
    """
    groups = {}
    for r in raw_rows:
        # Non-Shopify rows (simple schema) pass through unchanged
        if "jurisdiction_type" not in r:
            key = ("simple", id(r))
            groups[key] = r
            continue

        key = (
            r.get("order_number", ""),
            r.get("line_item_id", ""),
            r.get("date", ""),
            round(r.get("taxable_amount", 0), 2),
            round(r.get("gross_sales", 0), 2),
        )

        if key not in groups:
            # First row for this line item — clone and zero the tax (we'll sum below)
            merged = dict(r)
            merged["shopify_tax_amount"] = 0.0
            merged["shopify_tax_rate"] = 0.0
            merged["_state_tax_collected"] = 0.0
            merged["_county_tax_collected"] = 0.0
            merged["_jurisdictions_seen"] = []
            groups[key] = merged

        agg = groups[key]
        agg["shopify_tax_amount"] += r.get("shopify_tax_amount", 0)
        agg["_jurisdictions_seen"].append(r.get("jurisdiction_type", ""))

        if r.get("jurisdiction_type") == "state":
            agg["_state_tax_collected"] += r.get("shopify_tax_amount", 0)
        elif r.get("jurisdiction_type") == "county":
            agg["_county_tax_collected"] += r.get("shopify_tax_amount", 0)
            # If we learned the county from the county row, use it
            if r.get("delivery_county") and not agg.get("delivery_county"):
                agg["delivery_county"] = r["delivery_county"]

        # Effective rate = sum / taxable
        if agg.get("taxable_amount"):
            agg["shopify_tax_rate"] = round(
                agg["shopify_tax_amount"] / agg["taxable_amount"], 4
            )

    # Return in original order
    return list(groups.values())


def load_excel_file(path: Path) -> List[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = _normalize_headers(header_row)
    txns = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = dict(zip(headers, row))
        t = _row_to_txn(rec, path.name)
        if t:
            txns.append(t)
    return txns


def load_csv_file(path: Path) -> List[dict]:
    import csv
    txns = []
    # Try tab-separated first (your paste used tabs), then comma
    for delim in ["\t", ","]:
        try:
            with path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f, delimiter=delim)
                first_row = next(reader, None)
                if first_row is None:
                    continue
                # If only one column was parsed, wrong delimiter — try next
                if len(first_row.keys()) == 1 and delim == "\t":
                    continue
                rec0 = {(k or "").strip().lower(): v for k, v in first_row.items()}
                t = _row_to_txn(rec0, path.name)
                if t:
                    txns.append(t)
                for r in reader:
                    rec = {(k or "").strip().lower(): v for k, v in r.items()}
                    t = _row_to_txn(rec, path.name)
                    if t:
                        txns.append(t)
            break  # succeeded
        except Exception:
            continue
    return txns


def load_all_daily_data(
    folder: str,
    period_start: Optional[str] = None,
    period_end: Optional[str] = None,
) -> Tuple[List[dict], List[str]]:
    folder_path = Path(folder)
    all_txns: List[dict] = []
    files_read: List[str] = []

    for f in sorted(folder_path.iterdir()):
        if f.name.startswith(".") or f.name.startswith("_") or f.name.startswith("~$"):
            continue
        suffix = f.suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                loaded = load_excel_file(f)
            elif suffix in {".csv", ".tsv", ".txt"}:
                loaded = load_csv_file(f)
            else:
                continue
            all_txns.extend(loaded)
            files_read.append(f"{f.name} ({len(loaded)} rows)")
        except Exception as e:
            print(f"  ! Skipped {f.name}: {e}")

    # Date filter
    if period_start or period_end:
        ps = datetime.fromisoformat(period_start) if period_start else None
        pe = datetime.fromisoformat(period_end) if period_end else None
        filtered = []
        for t in all_txns:
            if not t.get("date"):
                filtered.append(t)
                continue
            try:
                d = datetime.fromisoformat(t["date"])
            except ValueError:
                filtered.append(t)
                continue
            if ps and d < ps:
                continue
            if pe and d > pe:
                continue
            filtered.append(t)
        all_txns = filtered

    # Collapse Shopify's State+County duplicate rows into unified transactions
    raw_count = len(all_txns)
    all_txns = _aggregate_shopify_rows(all_txns)
    aggregated_count = len(all_txns)
    if raw_count != aggregated_count:
        files_read.append(
            f"(aggregated {raw_count} Shopify rows → {aggregated_count} line items)"
        )

    return all_txns, files_read


def load_per_file(
    folder: str,
    period_start: Optional[str] = None,
    period_end: Optional[str] = None,
) -> List[Tuple[str, List[dict]]]:
    """Load each sales file separately, keeping them isolated.

    Returns a list of (filename, transactions) tuples — one per input file.
    This is what you want when each file represents a separate filing
    period (e.g. one per month or quarter).
    """
    folder_path = Path(folder)
    results: List[Tuple[str, List[dict]]] = []

    for f in sorted(folder_path.iterdir()):
        if f.name.startswith(".") or f.name.startswith("_") or f.name.startswith("~$"):
            continue
        suffix = f.suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                loaded = load_excel_file(f)
            elif suffix in {".csv", ".tsv", ".txt"}:
                loaded = load_csv_file(f)
            else:
                continue
        except Exception as e:
            print(f"  ! Skipped {f.name}: {e}")
            continue

        # Date filter (optional)
        if period_start or period_end:
            ps = datetime.fromisoformat(period_start) if period_start else None
            pe = datetime.fromisoformat(period_end) if period_end else None
            filtered = []
            for t in loaded:
                if not t.get("date"):
                    filtered.append(t)
                    continue
                try:
                    d = datetime.fromisoformat(t["date"])
                except ValueError:
                    filtered.append(t)
                    continue
                if ps and d < ps:
                    continue
                if pe and d > pe:
                    continue
                filtered.append(t)
            loaded = filtered

        # Aggregate State+County pairs within this file only
        aggregated = _aggregate_shopify_rows(loaded)

        if aggregated:
            results.append((f.name, aggregated))

    return results
