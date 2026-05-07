"""
Generates the Excel report of downloaded bills.
"""

from pathlib import Path
from datetime import datetime
from typing import List, Dict
import logging

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Creates a formatted Excel report of all downloaded bills."""

    HEADER_BG = "1F4E79"       # Dark blue
    HEADER_FONT = "FFFFFF"     # White
    SUCCESS_BG = "E2EFDA"      # Light green
    FAIL_BG = "FCE4D6"         # Light red/orange
    ALT_ROW = "F2F7FF"         # Light blue alternating row
    SUMMARY_BG = "FFF2CC"      # Yellow summary

    def __init__(self, download_dir: Path, logger: logging.Logger):
        self.download_dir = download_dir
        self.logger = logger

    def generate(self, results: List[Dict]) -> Path:
        """Generate and save the Excel report. Returns path to the file."""
        date_str = datetime.now().strftime("%Y-%m-%d")
        report_path = self.download_dir / f"Bills_Report_{date_str}.xlsx"

        wb = Workbook()

        # Sheet 1: All Bills Detail
        ws_detail = wb.active
        ws_detail.title = "Downloaded Bills"
        self._build_detail_sheet(ws_detail, results)

        # Sheet 2: Summary by Batch
        ws_summary = wb.create_sheet("Batch Summary")
        self._build_summary_sheet(ws_summary, results)

        # Sheet 3: Failed Downloads
        failed = [r for r in results if not r.get("success")]
        if failed:
            ws_failed = wb.create_sheet("Failed Downloads")
            self._build_failed_sheet(ws_failed, failed)

        wb.save(report_path)
        self.logger.info(f"Excel report saved: {report_path}")
        return report_path

    def _build_detail_sheet(self, ws, results: List[Dict]):
        """Build the main detail sheet."""
        headers = [
            "Sr No.", "Bill Number", "Bill Date", "Batch #",
            "Position in Batch", "Downloaded On", "Downloaded At",
            "File Name", "File Path", "Status"
        ]
        col_widths = [8, 18, 14, 10, 18, 16, 14, 30, 50, 12]

        # Title row
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = f"XPO Billing Download Report — {datetime.now().strftime('%B %d, %Y')}"
        title_cell.font = Font(name="Arial", size=14, bold=True, color=self.HEADER_FONT)
        title_cell.fill = PatternFill("solid", fgColor=self.HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Header row
        header_fill = PatternFill("solid", fgColor="2E74B5")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = 32

        # Freeze header rows
        ws.freeze_panes = "A3"

        # Data rows
        for idx, result in enumerate(results, start=1):
            row = idx + 2
            is_success = result.get("success", False)
            row_fill_color = self.SUCCESS_BG if is_success else self.FAIL_BG
            if is_success and idx % 2 == 0:
                row_fill_color = self.ALT_ROW

            fill = PatternFill("solid", fgColor=row_fill_color)
            row_data = [
                idx,
                result.get("bill_number", ""),
                result.get("bill_date", ""),
                result.get("batch_num", ""),
                result.get("position_in_batch", ""),
                result.get("downloaded_on", ""),
                result.get("downloaded_at", ""),
                result.get("file_name", ""),
                result.get("file_path", ""),
                "✓ Success" if is_success else "✗ Failed",
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = self._thin_border()
                if col_idx == 10:  # Status column
                    cell.font = Font(
                        name="Arial", size=9, bold=True,
                        color="375623" if is_success else "C00000"
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row].height = 18

        # Add totals row
        total_row = len(results) + 3
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
        ws.cell(row=total_row, column=9, value=f"=COUNTA(J3:J{total_row-1})")
        success_count = sum(1 for r in results if r.get("success"))
        failed_count = len(results) - success_count

        summary_fill = PatternFill("solid", fgColor=self.SUMMARY_BG)
        for col in range(1, 11):
            ws.cell(row=total_row, column=col).fill = summary_fill

        ws.cell(row=total_row, column=8, value=f"Successful: {success_count}  |  Failed: {failed_count}")
        ws.cell(row=total_row, column=8).font = Font(bold=True, name="Arial", size=10)

    def _build_summary_sheet(self, ws, results: List[Dict]):
        """Build batch summary sheet."""
        # Header
        ws.merge_cells("A1:E1")
        ws["A1"].value = "Batch Download Summary"
        ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=self.HEADER_BG)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        headers = ["Batch #", "Bills in Batch", "Successful", "Failed", "Downloaded On"]
        widths = [12, 16, 14, 12, 20]
        header_fill = PatternFill("solid", fgColor="2E74B5")

        for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Group results by batch
        batches = {}
        for r in results:
            b = r.get("batch_num", 1)
            if b not in batches:
                batches[b] = []
            batches[b].append(r)

        for row_idx, (batch_num, batch_results) in enumerate(sorted(batches.items()), start=3):
            success = sum(1 for r in batch_results if r.get("success"))
            failed = len(batch_results) - success
            date = batch_results[0].get("downloaded_on", "")

            fill = PatternFill("solid", fgColor=self.ALT_ROW if row_idx % 2 == 0 else "FFFFFF")
            row_data = [batch_num, len(batch_results), success, failed, date]

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = self._thin_border()

        # Grand totals
        total_row = len(batches) + 3
        total_data = [
            "GRAND TOTAL",
            len(results),
            sum(1 for r in results if r.get("success")),
            sum(1 for r in results if not r.get("success")),
            ""
        ]
        total_fill = PatternFill("solid", fgColor=self.SUMMARY_BG)
        for col_idx, val in enumerate(total_data, start=1):
            cell = ws.cell(row=total_row, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._thin_border()

    def _build_failed_sheet(self, ws, failed: List[Dict]):
        """Build failed downloads sheet."""
        ws.merge_cells("A1:D1")
        ws["A1"].value = "Failed Downloads - Manual Action Required"
        ws["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="C00000")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Bill Number", "Bill Date", "Batch #", "Raw Info"]
        widths = [20, 16, 12, 50]
        for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.border = self._thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        for row_idx, r in enumerate(failed, start=3):
            row_data = [r.get("bill_number",""), r.get("bill_date",""), r.get("batch_num",""), r.get("raw_text","")]
            fill = PatternFill("solid", fgColor=self.FAIL_BG)
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill
                cell.border = self._thin_border()

    def _thin_border(self):
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)
